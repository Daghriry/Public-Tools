import os
import pandas as pd
from pathlib import Path
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
import humanize
from collections import defaultdict
import time
import hashlib

class DiskAnalyzer:
    def __init__(self, root_path):
        self.root_path = Path(root_path)
        self.files_data = []
        self.folders_data = []
        self.file_types_summary = defaultdict(lambda: {'count': 0, 'size': 0})
        self.duplicates = defaultdict(list)  # hash -> list of file paths
        self.errors = []
        
    def format_bytes(self, bytes_value):
        """Convert bytes to human readable format"""
        return humanize.naturalsize(bytes_value, binary=True)
    
    def get_file_hash(self, file_path, quick=True):
        """Calculate file hash for duplicate detection"""
        try:
            # For quick mode, only hash first 1MB for large files
            hash_md5 = hashlib.md5()
            path_str = str(file_path)
            
            # Handle long paths in Windows
            if os.name == 'nt' and len(path_str) > 260 and not path_str.startswith('\\\\?\\'):
                path_str = '\\\\?\\' + os.path.abspath(path_str)
            
            with open(path_str, "rb") as f:
                if quick and os.path.getsize(path_str) > 1024*1024:
                    # For files > 1MB, only hash first 1MB for speed
                    hash_md5.update(f.read(1024*1024))
                else:
                    # For smaller files, hash entire file
                    for chunk in iter(lambda: f.read(4096), b""):
                        hash_md5.update(chunk)
            
            return hash_md5.hexdigest()
        except:
            return None
    
    def get_file_info(self, file_path):
        """Get detailed information about a file"""
        try:
            # Handle long paths in Windows
            path_str = str(file_path)
            if os.name == 'nt' and len(path_str) > 260 and not path_str.startswith('\\\\?\\'):
                path_str = '\\\\?\\' + os.path.abspath(path_str)
            
            stat = os.stat(path_str)
            
            # Get relative path safely
            try:
                rel_path = file_path.relative_to(self.root_path)
                depth = len(rel_path.parts)
            except:
                depth = 0
            
            file_info = {
                'path': str(file_path),
                'name': file_path.name,
                'parent_folder': str(file_path.parent),
                'extension': file_path.suffix.lower() if file_path.suffix else 'no_extension',
                'size_bytes': stat.st_size,
                'size_readable': self.format_bytes(stat.st_size),
                'modified_time': datetime.datetime.fromtimestamp(stat.st_mtime),
                'created_time': datetime.datetime.fromtimestamp(stat.st_ctime),
                'depth': depth
            }
            
            # Calculate hash for duplicate detection (only for files < 100MB for performance)
            if stat.st_size < 100*1024*1024:
                file_hash = self.get_file_hash(file_path)
                if file_hash:
                    file_info['hash'] = file_hash
                    self.duplicates[file_hash].append({
                        'path': str(file_path),
                        'name': file_path.name,
                        'size': stat.st_size
                    })
            
            return file_info
        except Exception as e:
            self.errors.append(f"Error reading file {file_path}: {str(e)}")
            return None
    
    def calculate_folder_size(self, folder_path):
        """Calculate total size of a folder including all subfolders"""
        total_size = 0
        file_count = 0
        subfolder_count = 0
        
        try:
            # Use os.walk for better error handling
            for root, dirs, files in os.walk(folder_path):
                # Count subfolders
                subfolder_count += len(dirs)
                
                # Count and sum files
                for file in files:
                    file_count += 1
                    try:
                        file_path = os.path.join(root, file)
                        # Handle long paths in Windows
                        if os.name == 'nt' and len(file_path) > 260:
                            file_path = '\\\\?\\' + os.path.abspath(file_path)
                        total_size += os.path.getsize(file_path)
                    except:
                        # Skip files that can't be accessed
                        pass
        except Exception as e:
            self.errors.append(f"Error calculating folder size for {folder_path}: {str(e)}")
        
        return total_size, file_count, subfolder_count
    
    def safe_walk(self, path):
        """Safely walk through directory tree, handling long paths and errors"""
        try:
            # For Windows, try to use extended path format for long paths
            if os.name == 'nt' and len(str(path)) > 200:
                path_str = str(path)
                if not path_str.startswith('\\\\?\\'):
                    path_str = '\\\\?\\' + os.path.abspath(path_str)
                    path = Path(path_str)
            
            for item in os.scandir(path):
                yield Path(item.path)
                if item.is_dir(follow_symlinks=False):
                    try:
                        yield from self.safe_walk(item.path)
                    except Exception as e:
                        self.errors.append(f"Cannot access folder: {item.path} - {str(e)}")
        except Exception as e:
            self.errors.append(f"Error scanning: {path} - {str(e)}")
    
    def analyze_directory(self):
        """Analyze the directory structure and collect all data"""
        print(f"Starting analysis of: {self.root_path}")
        start_time = time.time()
        
        # Progress tracking
        file_count = 0
        folder_count = 0
        last_update = time.time()
        
        # Collect all files and folders in one pass
        print("Scanning files and folders...")
        print("(Progress updates every 5 seconds)")
        
        try:
            for item_path in self.safe_walk(self.root_path):
                try:
                    # Update progress every 5 seconds
                    if time.time() - last_update > 5:
                        print(f"  Progress: {file_count:,} files, {folder_count:,} folders scanned...")
                        last_update = time.time()
                    
                    if item_path.is_file():
                        file_count += 1
                        file_info = self.get_file_info(item_path)
                        if file_info:
                            self.files_data.append(file_info)
                            # Update file type summary
                            ext = file_info['extension']
                            self.file_types_summary[ext]['count'] += 1
                            self.file_types_summary[ext]['size'] += file_info['size_bytes']
                    
                    elif item_path.is_dir():
                        folder_count += 1
                        try:
                            size, files_in_folder, subfolders_in_folder = self.calculate_folder_size(item_path)
                            rel_path = item_path.relative_to(self.root_path) if item_path != self.root_path else Path('.')
                            self.folders_data.append({
                                'path': str(item_path),
                                'name': item_path.name,
                                'parent_folder': str(item_path.parent),
                                'size_bytes': size,
                                'size_readable': self.format_bytes(size),
                                'file_count': files_in_folder,
                                'subfolder_count': subfolders_in_folder,
                                'depth': len(rel_path.parts) if rel_path != Path('.') else 0,
                                'total_items': files_in_folder + subfolders_in_folder
                            })
                        except Exception as e:
                            self.errors.append(f"Error processing folder {item_path}: {str(e)}")
                
                except Exception as e:
                    self.errors.append(f"Error processing item {item_path}: {str(e)}")
        
        except Exception as e:
            self.errors.append(f"Critical error during scan: {str(e)}")
        
        # Add root folder if not already added
        if not any(Path(f['path']) == self.root_path for f in self.folders_data):
            try:
                root_size, root_files, root_subfolders = self.calculate_folder_size(self.root_path)
                self.folders_data.append({
                    'path': str(self.root_path),
                    'name': self.root_path.name,
                    'parent_folder': str(self.root_path.parent),
                    'size_bytes': root_size,
                    'size_readable': self.format_bytes(root_size),
                    'file_count': root_files,
                    'subfolder_count': root_subfolders,
                    'depth': 0,
                    'total_items': root_files + root_subfolders
                })
            except Exception as e:
                self.errors.append(f"Error calculating root folder size: {str(e)}")
        
        elapsed_time = time.time() - start_time
        print(f"\nAnalysis completed in {elapsed_time:.2f} seconds")
        print(f"Found {len(self.files_data):,} files and {len(self.folders_data):,} folders")
        if self.errors:
            print(f"Encountered {len(self.errors)} errors (will be listed in the report)")
    
    def get_size_color(self, size_bytes):
        """Return color based on file/folder size"""
        if size_bytes >= 1024**3:  # >= 1GB - Red
            return 'FF0000'
        elif size_bytes >= 500*1024**2:  # >= 500MB - Orange
            return 'FFA500'
        elif size_bytes >= 100*1024**2:  # >= 100MB - Yellow
            return 'FFFF00'
        elif size_bytes >= 10*1024**2:  # >= 10MB - Light Green
            return '90EE90'
        else:  # < 10MB - Light Blue
            return 'ADD8E6'
    
    def create_excel_report(self, output_file='disk_analysis_report.xlsx'):
        """Create comprehensive Excel report with multiple sheets"""
        print(f"Creating Excel report: {output_file}")
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        self.create_summary_sheet(ws_summary)
        
        # 2. Charts Sheet
        ws_charts = wb.create_sheet("Charts")
        self.create_charts_sheet(ws_charts)
        
        # 3. Folders Sheet
        ws_folders = wb.create_sheet("Folders")
        self.create_folders_sheet(ws_folders)
        
        # 4. Files Sheet
        ws_files = wb.create_sheet("Files")
        self.create_files_sheet(ws_files)
        
        # 5. Duplicate Files Sheet
        ws_duplicates = wb.create_sheet("Duplicate Files")
        self.create_duplicates_sheet(ws_duplicates)
        
        # 6. File Types Sheet
        ws_types = wb.create_sheet("File Types")
        self.create_file_types_sheet(ws_types)
        
        # 7. Large Files Sheet (Top 100)
        ws_large = wb.create_sheet("Large Files")
        self.create_large_files_sheet(ws_large)
        
        # 8. Errors Sheet (if any)
        if self.errors:
            ws_errors = wb.create_sheet("Errors")
            self.create_errors_sheet(ws_errors)
        
        # Save the workbook
        wb.save(output_file)
        print(f"Report saved successfully: {output_file}")
    
    def create_summary_sheet(self, ws):
        """Create summary statistics sheet"""
        # Title
        ws['A1'] = 'Disk Space Analysis Report'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Analysis info
        ws['A3'] = 'Analysis Date:'
        ws['B3'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = 'Root Path:'
        ws['B4'] = str(self.root_path)
        
        # Statistics
        total_size = sum(f['size_bytes'] for f in self.files_data)
        duplicate_count = sum(1 for files in self.duplicates.values() if len(files) > 1)
        
        ws['A6'] = 'Total Files:'
        ws['B6'] = len(self.files_data)
        ws['A7'] = 'Total Folders:'
        ws['B7'] = len(self.folders_data)
        ws['A8'] = 'Total Size:'
        ws['B8'] = self.format_bytes(total_size)
        ws['A9'] = 'Average File Size:'
        ws['B9'] = self.format_bytes(total_size / len(self.files_data)) if self.files_data else '0 B'
        ws['A10'] = 'Duplicate File Groups:'
        ws['B10'] = duplicate_count
        ws['B10'].font = Font(color='FF0000' if duplicate_count > 0 else '000000')
        
        # Top 10 largest folders
        ws['A12'] = 'Top 10 Largest Folders'
        ws['A12'].font = Font(bold=True, size=14)
        
        headers = ['Rank', 'Folder Path', 'Size', 'Files']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=13, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        sorted_folders = sorted(self.folders_data, key=lambda x: x['size_bytes'], reverse=True)[:10]
        for idx, folder in enumerate(sorted_folders, 1):
            ws.cell(row=13+idx, column=1, value=idx)
            ws.cell(row=13+idx, column=2, value=folder['path'])
            ws.cell(row=13+idx, column=3, value=folder['size_readable'])
            ws.cell(row=13+idx, column=4, value=folder['file_count'])
            
            # Apply color based on size
            color = self.get_size_color(folder['size_bytes'])
            for col in range(1, 5):
                ws.cell(row=13+idx, column=col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type='solid'
                )
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
    
    def create_charts_sheet(self, ws):
        """Create visual charts sheet"""
        ws['A1'] = 'Disk Space Analysis Charts'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Prepare data for charts
        # 1. Top 10 File Types by Size
        ws['A3'] = 'Top 10 File Types by Size'
        ws['A3'].font = Font(bold=True, size=14)
        
        # Sort file types by size
        types_sorted = sorted(self.file_types_summary.items(), 
                            key=lambda x: x[1]['size'], reverse=True)[:10]
        
        # Add headers
        ws['A5'] = 'Extension'
        ws['B5'] = 'Size (MB)'
        ws['C5'] = 'Count'
        ws['A5'].font = Font(bold=True)
        ws['B5'].font = Font(bold=True)
        ws['C5'].font = Font(bold=True)
        
        # Add data
        for idx, (ext, data) in enumerate(types_sorted, 6):
            ws[f'A{idx}'] = ext
            ws[f'B{idx}'] = round(data['size'] / (1024*1024), 2)  # Convert to MB
            ws[f'C{idx}'] = data['count']
        
        # Create Pie Chart for File Types
        pie = PieChart()
        pie.title = "Storage Distribution by File Type"
        pie.style = 10
        pie.width = 15
        pie.height = 10
        
        # Data for pie chart
        data = Reference(ws, min_col=2, min_row=5, max_row=5+len(types_sorted))
        categories = Reference(ws, min_col=1, min_row=6, max_row=5+len(types_sorted))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(categories)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        
        ws.add_chart(pie, "E3")
        
        # 2. Top 10 Folders by Size
        ws['A20'] = 'Top 10 Folders by Size'
        ws['A20'].font = Font(bold=True, size=14)
        
        # Headers
        ws['A22'] = 'Folder Name'
        ws['B22'] = 'Size (GB)'
        ws['A22'].font = Font(bold=True)
        ws['B22'].font = Font(bold=True)
        
        # Add folder data
        folders_sorted = sorted(self.folders_data, key=lambda x: x['size_bytes'], reverse=True)[:10]
        for idx, folder in enumerate(folders_sorted, 23):
            folder_name = folder['name'] if folder['name'] else 'Root'
            ws[f'A{idx}'] = folder_name[:30] + '...' if len(folder_name) > 30 else folder_name
            ws[f'B{idx}'] = round(folder['size_bytes'] / (1024**3), 2)  # Convert to GB
        
        # Create Bar Chart for Folders
        bar = BarChart()
        bar.title = "Top 10 Folders by Size (GB)"
        bar.style = 10
        bar.width = 15
        bar.height = 10
        bar.x_axis.title = "Folders"
        bar.y_axis.title = "Size (GB)"
        
        data = Reference(ws, min_col=2, min_row=22, max_row=22+len(folders_sorted))
        categories = Reference(ws, min_col=1, min_row=23, max_row=22+len(folders_sorted))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(categories)
        
        ws.add_chart(bar, "E20")
        
        # 3. Storage Summary
        ws['A40'] = 'Storage Summary'
        ws['A40'].font = Font(bold=True, size=14)
        
        total_size = sum(f['size_bytes'] for f in self.files_data)
        
        # Calculate size by category
        categories = {
            'Documents': ['.doc', '.docx', '.pdf', '.txt', '.odt', '.rtf', '.tex'],
            'Images': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.svg', '.ico', '.tiff'],
            'Videos': ['.mp4', '.avi', '.mkv', '.mov', '.wmv', '.flv', '.webm'],
            'Audio': ['.mp3', '.wav', '.flac', '.aac', '.ogg', '.wma', '.m4a'],
            'Archives': ['.zip', '.rar', '.7z', '.tar', '.gz', '.bz2', '.xz'],
            'Code': ['.py', '.js', '.java', '.cpp', '.c', '.cs', '.php', '.rb', '.go'],
            'Other': []
        }
        
        category_sizes = defaultdict(int)
        for ext, data in self.file_types_summary.items():
            categorized = False
            for category, extensions in categories.items():
                if ext in extensions:
                    category_sizes[category] += data['size']
                    categorized = True
                    break
            if not categorized:
                category_sizes['Other'] += data['size']
        
        # Add category data
        ws['A42'] = 'Category'
        ws['B42'] = 'Size (GB)'
        ws['A42'].font = Font(bold=True)
        ws['B42'].font = Font(bold=True)
        
        row = 43
        for category, size in sorted(category_sizes.items(), key=lambda x: x[1], reverse=True):
            if size > 0:
                ws[f'A{row}'] = category
                ws[f'B{row}'] = round(size / (1024**3), 2)
                row += 1
        
        # Create Pie Chart for Categories
        pie2 = PieChart()
        pie2.title = "Storage by Category"
        pie2.style = 10
        pie2.width = 15
        pie2.height = 10
        
        data = Reference(ws, min_col=2, min_row=42, max_row=row-1)
        categories = Reference(ws, min_col=1, min_row=43, max_row=row-1)
        pie2.add_data(data, titles_from_data=True)
        pie2.set_categories(categories)
        pie2.dataLabels = DataLabelList()
        pie2.dataLabels.showPercent = True
        
        ws.add_chart(pie2, "E40")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def create_duplicates_sheet(self, ws):
        """Create duplicate files sheet"""
        ws['A1'] = 'Duplicate Files Analysis'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Filter out duplicates (files with same hash)
        duplicate_groups = {h: files for h, files in self.duplicates.items() if len(files) > 1}
        
        if not duplicate_groups:
            ws['A3'] = 'No duplicate files found!'
            ws['A3'].font = Font(size=14, color='008000')
            return
        
        # Calculate potential savings
        total_wasted = 0
        for hash_val, files in duplicate_groups.items():
            if len(files) > 1:
                # Space wasted = size * (count - 1)
                total_wasted += files[0]['size'] * (len(files) - 1)
        
        ws['A3'] = 'Summary:'
        ws['A4'] = f'Total duplicate groups: {len(duplicate_groups)}'
        ws['A5'] = f'Total wasted space: {self.format_bytes(total_wasted)}'
        ws['A5'].font = Font(bold=True, color='FF0000')
        
        # Headers
        headers = ['Group', 'File Name', 'Path', 'Size', 'Modified Date', 'Action']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Add duplicate data
        row = 9
        group_num = 1
        
        for hash_val, files in sorted(duplicate_groups.items(), 
                                     key=lambda x: x[1][0]['size'], reverse=True):
            # Sort files in group by path
            files_sorted = sorted(files, key=lambda x: x['path'])
            
            first_file = True
            for file in files_sorted:
                ws.cell(row=row, column=1, value=f'Group {group_num}')
                ws.cell(row=row, column=2, value=file['name'])
                ws.cell(row=row, column=3, value=file['path'])
                ws.cell(row=row, column=4, value=self.format_bytes(file['size']))
                
                # Get file modification date
                try:
                    mtime = os.path.getmtime(file['path'])
                    ws.cell(row=row, column=5, value=datetime.datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M'))
                except:
                    ws.cell(row=row, column=5, value='Unknown')
                
                # Suggest action
                if first_file:
                    ws.cell(row=row, column=6, value='Keep (Original)')
                    ws.cell(row=row, column=6).font = Font(color='008000')
                    first_file = False
                else:
                    ws.cell(row=row, column=6, value='Consider Deleting')
                    ws.cell(row=row, column=6).font = Font(color='FF0000')
                
                # Apply alternating colors for groups
                if group_num % 2 == 0:
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color='F0F0F0', end_color='F0F0F0', fill_type='solid'
                        )
                
                row += 1
            
            # Add separator between groups
            row += 1
            group_num += 1
            
            # Limit to first 1000 groups for performance
            if group_num > 1000:
                ws.cell(row=row, column=1, value=f"Showing first 1000 duplicate groups out of {len(duplicate_groups)} total")
                ws.cell(row=row, column=1).font = Font(italic=True, color='FF0000')
                break
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        # Add autofilter
        ws.auto_filter.ref = f"A8:F{row-1}"
    
    def create_folders_sheet(self, ws):
        """Create detailed folders sheet"""
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(self.folders_data)
        df = df.sort_values('size_bytes', ascending=False)
        
        # Headers
        headers = ['Folder Path', 'Name', 'Size (Readable)', 'Size (Bytes)', 
                   'Files', 'Subfolders', 'Total Items', 'Depth', 'Parent Folder']
        
        # Add headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for idx, row in df.iterrows():
            row_num = len(ws['A']) + 1
            ws.cell(row=row_num, column=1, value=row['path'])
            ws.cell(row=row_num, column=2, value=row['name'])
            ws.cell(row=row_num, column=3, value=row['size_readable'])
            ws.cell(row=row_num, column=4, value=row['size_bytes'])
            ws.cell(row=row_num, column=5, value=row['file_count'])
            ws.cell(row=row_num, column=6, value=row['subfolder_count'])
            ws.cell(row=row_num, column=7, value=row['total_items'])
            ws.cell(row=row_num, column=8, value=row['depth'])
            ws.cell(row=row_num, column=9, value=row['parent_folder'])
            
            # Apply color based on size
            color = self.get_size_color(row['size_bytes'])
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type='solid'
                )
        
        # Adjust column widths
        column_widths = [60, 30, 15, 15, 10, 12, 12, 8, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add autofilter
        ws.auto_filter.ref = f"A1:I{len(df) + 1}"
    
    def create_files_sheet(self, ws):
        """Create detailed files sheet"""
        # Convert to DataFrame
        df = pd.DataFrame(self.files_data)
        df = df.sort_values('size_bytes', ascending=False)
        
        # Headers
        headers = ['File Path', 'Name', 'Extension', 'Size (Readable)', 'Size (Bytes)', 
                   'Modified', 'Created', 'Depth', 'Parent Folder']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add data (limit to 10000 rows for performance)
        for idx, row in df.head(10000).iterrows():
            row_num = len(ws['A']) + 1
            ws.cell(row=row_num, column=1, value=row['path'])
            ws.cell(row=row_num, column=2, value=row['name'])
            ws.cell(row=row_num, column=3, value=row['extension'])
            ws.cell(row=row_num, column=4, value=row['size_readable'])
            ws.cell(row=row_num, column=5, value=row['size_bytes'])
            ws.cell(row=row_num, column=6, value=row['modified_time'].strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=7, value=row['created_time'].strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=8, value=row['depth'])
            ws.cell(row=row_num, column=9, value=row['parent_folder'])
            
            # Apply color
            color = self.get_size_color(row['size_bytes'])
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type='solid'
                )
        
        # Note if limited
        if len(df) > 10000:
            row_num = len(ws['A']) + 2
            ws.cell(row=row_num, column=1, value=f"Note: Showing only top 10,000 files out of {len(df)} total files")
            ws.cell(row=row_num, column=1).font = Font(italic=True, color='FF0000')
        
        # Adjust column widths
        column_widths = [60, 40, 10, 15, 15, 20, 20, 8, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add autofilter
        ws.auto_filter.ref = f"A1:I{min(len(df) + 1, 10001)}"
    
    def create_file_types_sheet(self, ws):
        """Create file types summary sheet"""
        # Prepare data
        types_data = []
        for ext, data in self.file_types_summary.items():
            types_data.append({
                'extension': ext,
                'count': data['count'],
                'total_size_bytes': data['size'],
                'total_size_readable': self.format_bytes(data['size']),
                'avg_size_bytes': data['size'] / data['count'] if data['count'] > 0 else 0,
                'avg_size_readable': self.format_bytes(data['size'] / data['count']) if data['count'] > 0 else '0 B'
            })
        
        df = pd.DataFrame(types_data)
        df = df.sort_values('total_size_bytes', ascending=False)
        
        # Headers
        headers = ['Extension', 'File Count', 'Total Size', 'Total Size (Bytes)', 
                   'Average Size', 'Average Size (Bytes)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for idx, row in df.iterrows():
            row_num = len(ws['A']) + 1
            ws.cell(row=row_num, column=1, value=row['extension'])
            ws.cell(row=row_num, column=2, value=row['count'])
            ws.cell(row=row_num, column=3, value=row['total_size_readable'])
            ws.cell(row=row_num, column=4, value=row['total_size_bytes'])
            ws.cell(row=row_num, column=5, value=row['avg_size_readable'])
            ws.cell(row=row_num, column=6, value=row['avg_size_bytes'])
            
            # Apply color based on total size
            color = self.get_size_color(row['total_size_bytes'])
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type='solid'
                )
        
        # Adjust column widths
        column_widths = [15, 15, 20, 20, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add autofilter
        ws.auto_filter.ref = f"A1:F{len(df) + 1}"
    
    def create_large_files_sheet(self, ws):
        """Create sheet with largest files"""
        # Get top 100 largest files
        df = pd.DataFrame(self.files_data)
        df = df.nlargest(100, 'size_bytes')
        
        # Title
        ws['A1'] = 'Top 100 Largest Files'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = ['Rank', 'File Name', 'Size', 'Size (Bytes)', 'Extension', 'Full Path']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for idx, row in df.iterrows():
            row_num = len(ws['A']) + 1
            ws.cell(row=row_num, column=1, value=row_num - 3)
            ws.cell(row=row_num, column=2, value=row['name'])
            ws.cell(row=row_num, column=3, value=row['size_readable'])
            ws.cell(row=row_num, column=4, value=row['size_bytes'])
            ws.cell(row=row_num, column=5, value=row['extension'])
            ws.cell(row=row_num, column=6, value=row['path'])
            
            # Apply color
            color = self.get_size_color(row['size_bytes'])
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type='solid'
                )
        
        # Adjust column widths
        column_widths = [8, 50, 15, 15, 10, 80]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_errors_sheet(self, ws):
        """Create sheet with errors encountered during analysis"""
        ws['A1'] = 'Errors Encountered During Analysis'
        ws['A1'].font = Font(bold=True, size=14, color='FF0000')
        
        ws['A3'] = 'Error Description'
        ws['A3'].font = Font(bold=True)
        
        for idx, error in enumerate(self.errors, 4):
            ws.cell(row=idx, column=1, value=error)
        
        ws.column_dimensions['A'].width = 100

def main():
    print("=== Disk Space Analyzer ===")
    print("This tool will analyze all files and folders in a specified path")
    print("and generate a detailed Excel report.\n")
    
    # Get path from user
    while True:
        path = input("Enter the path to analyze (or 'quit' to exit): ").strip()
        
        if path.lower() == 'quit':
            print("Exiting...")
            return
        
        # Remove quotes if present
        path = path.strip('"').strip("'")
        
        # Validate path
        if os.path.exists(path):
            break
        else:
            print(f"Error: Path '{path}' does not exist! Please try again.\n")
    
    # Generate output filename with timestamp in script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder_name = os.path.basename(path).replace(" ", "_").replace(":", "")
    output_file = os.path.join(script_dir, f"disk_analysis_{folder_name}_{timestamp}.xlsx")
    
    print(f"\nAnalyzing: {path}")
    print(f"Report will be saved to: {output_file}")
    print("\nThis may take several minutes depending on the number of files...")
    
    # Create analyzer and run analysis
    try:
        analyzer = DiskAnalyzer(path)
        analyzer.analyze_directory()
        analyzer.create_excel_report(output_file)
        
        # Print summary
        print("\n" + "="*50)
        print("ANALYSIS COMPLETE!")
        print("="*50)
        print(f"Total files analyzed: {len(analyzer.files_data):,}")
        print(f"Total folders analyzed: {len(analyzer.folders_data):,}")
        print(f"Errors encountered: {len(analyzer.errors)}")
        print(f"\nReport saved to: {output_file}")
        
        # Offer to open the file
        if os.name == 'nt':  # Windows
            open_file = input("\nDo you want to open the report now? (y/n): ").strip().lower()
            if open_file == 'y':
                os.startfile(output_file)
        
    except Exception as e:
        print(f"\nError during analysis: {str(e)}")
        return
    
    input("\nPress Enter to exit...")

if __name__ == "__main__":
    main()